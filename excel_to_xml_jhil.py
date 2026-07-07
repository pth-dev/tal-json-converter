import re
import io
import openpyxl
from datetime import datetime
from xml.etree import ElementTree as ET
from xml.dom import minidom

# ── Namespace constants ────────────────────────────────────────────────────────
NS      = "http://www.jhilburn.com/ns/mtm/orders"
XSI     = "http://www.w3.org/2001/XMLSchema-instance"
XSD     = "http://www.w3.org/2001/XMLSchema"
NIL_KEY = f"{{{XSI}}}nil"

# ── Fields that use xsi:nil="true" when empty ─────────────────────────────────
NIL_FIELDS = {
    "LowerBodyFabricNo",
    "LowerBody",
    "FirstShirtToCustomer",
    "SleevePlacket",
    "Button",
}

# ── Exact field order as in the original XML ──────────────────────────────────
FIELD_ORDER = [
    "OrderId", "DetailId", "PoDate", "PtnrNo",
    "CustomerFirstName", "CustomerLastName", "CustomerNumber", "OrderQty",
    "ChestBodyMeasurement", "BodyFit", "Style", "Fit", "SleeveFullness",
    "Collar", "Placket", "Sleeve", "PlacketButton", "Vent",
    "PocketLocation", "Pocket", "Button",
    "BaseFabricNo", "LowerBodyFabricNo", "ContrastFabric",
    "BtnSidePlacket", "InteriorCollarBand", "EntireCollar", "EntirePlacket",
    "DoubleStripes", "LowerBody", "ContrastThreadColor",
    "BtnThrContrast", "BtnHoleThrContrast",
    "Size", "Waist", "Hip",
    "SleeveRight", "SleeveLeft", "CuffRight", "CuffLeft",
    "Length", "Shoulder", "Armhole", "Elbow", "Neck",
    "Embroidery", "EmbroideryStyle",
    "MonogramInitial", "MonogramLocation", "MonogramColor", "MonogramStyle",
    "FOB", "Wash", "RetailPrice", "DirectToClient", "FirstShirtToCustomer",
    "ShipToName", "StreetAddress", "StreetAddressLine2",
    "City", "State", "ZipCode", "CustomerAddressCountry", "CustomerPhoneNumber",
    "Priority", "SleevePlacket", "ShoulderSlope", "ContrastPiping",
    "PocketOpening", "PlacketZipTape", "UtilityPocket",
    "CareInstruction", "FabricContent", "JHISKU", "SupporterInterlining",
    "ColorBlockFabricB", "ColorBlockFabricC", "GoldenStitchLabel",
]

# ── Numeric precision rules ────────────────────────────────────────────────────
PRECISION = {
    "Waist":       "%.2f",
    "Hip":         "%.2f",
    "SleeveRight": "%.6f",
    "SleeveLeft":  "%.6f",
    "CuffRight":   "%.4f",
    "CuffLeft":    "%.4f",
    "Length":      "%.2f",
    "Shoulder":    "%.6f",
    "Armhole":     "%.6f",
    "RetailPrice": "%.2f",
    "Neck":        "%.2f",
}

INT_FIELDS = {
    "OrderId", "DetailId", "PtnrNo", "CustomerNumber",
    "OrderQty", "ChestBodyMeasurement", "Size", "ShoulderSlope",
}

ZIPCODE_FIELD = "ZipCode"


def format_bodyfit(val):
    try:
        n = float(val)
        return f"+{int(n)}" if n >= 0 else str(int(n))
    except (TypeError, ValueError):
        return str(val) if val is not None else "+0"


def format_date(val):
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%dT%H:%M:%S")
    s = str(val).strip()
    if re.match(r"\d{4}-\d{2}-\d{2}T\d{2}:\d{2}:\d{2}", s):
        return s
    for fmt in ("%m/%d/%y %H:%M", "%m/%d/%Y %H:%M", "%m/%d/%y", "%m/%d/%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%dT%H:%M:%S")
        except ValueError:
            continue
    return s


def format_numeric(field, val):
    if val is None or str(val).strip() == "":
        return None
    try:
        n = float(str(val).replace(",", ""))
        if field in PRECISION:
            return PRECISION[field] % n
        elif field in INT_FIELDS:
            return str(int(n))
        else:
            return str(int(n)) if n == int(n) else str(n)
    except (TypeError, ValueError):
        return str(val)


def read_excel_bytes(excel_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        record = {headers[i]: val for i, val in enumerate(row) if i < len(headers) and headers[i]}
        rows.append(record)

    return rows


def build_knits_order(record):
    order = ET.Element(f"{{{NS}}}KnitsOrder")

    for field in FIELD_ORDER:
        raw = record.get(field)

        if field == "PoDate":
            val_str = format_date(raw)

        elif field == "BodyFit":
            val_str = format_bodyfit(raw)

        elif field == ZIPCODE_FIELD:
            if raw is None:
                val_str = None
            elif isinstance(raw, (int, float)):
                val_str = str(int(raw)).zfill(5)
            else:
                val_str = str(raw).strip() or None

        elif field in INT_FIELDS:
            val_str = format_numeric(field, raw)

        elif field in PRECISION:
            val_str = format_numeric(field, raw)

        elif field in ("Size", "Neck", "ChestBodyMeasurement", "ShoulderSlope", "PlacketButton"):
            if raw is None:
                val_str = None
            else:
                s = str(raw).strip()
                try:
                    n = float(s)
                    if field == "Neck":
                        val_str = "%.2f" % n
                    elif field in ("ShoulderSlope", "ChestBodyMeasurement"):
                        val_str = str(int(n))
                    elif field == "Size":
                        val_str = str(int(n)) if n == int(n) else str(n)
                    else:
                        val_str = s
                except ValueError:
                    val_str = s if s else None

        else:
            if raw is None:
                val_str = None
            elif isinstance(raw, float) and raw == int(raw):
                val_str = str(int(raw))
            elif isinstance(raw, (int, float)):
                val_str = str(raw)
            else:
                val_str = str(raw).strip() or None

        el = ET.SubElement(order, f"{{{NS}}}{field}")
        if val_str is None or val_str == "":
            if field in NIL_FIELDS:
                el.set(NIL_KEY, "true")
        else:
            el.text = val_str

    return order


def excel_to_xml(excel_bytes: bytes) -> str:
    ET.register_namespace("",    NS)
    ET.register_namespace("xsi", XSI)
    ET.register_namespace("xsd", XSD)

    records = read_excel_bytes(excel_bytes)
    root = ET.Element(f"{{{NS}}}KnitsOrders")

    errors = []
    for i, record in enumerate(records, 1):
        try:
            root.append(build_knits_order(record))
        except Exception as e:
            errors.append(f"Record {i} (OrderId={record.get('OrderId')}): {e}")

    raw = ET.tostring(root, encoding="unicode", xml_declaration=False)
    dom = minidom.parseString(f'<?xml version="1.0" encoding="utf-8"?>{raw}')
    lines = dom.toprettyxml(indent="  ", encoding=None).splitlines()
    if lines[0].startswith("<?xml"):
        lines = lines[1:]
    xml_str = '<?xml version="1.0" encoding="utf-8"?>\n' + "\n".join(lines)

    return xml_str, len(records), errors
