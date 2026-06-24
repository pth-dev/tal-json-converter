import json
import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

HEADER_COLS_COMMON = [
    "OrderNumber", "ReferenceNo", "RSVS", "OrderDate", "OrderType",
    "ShipmentMethod", "CustomerName",
    "ShiptoCustomerName", "CustomerAddress1", "CustomerAddress2", "CustomerAddress3",
    "CustomerAddressCity", "CustomerAddressState", "CustomerAddressZip",
    "CustomerAddressCountry", "CustomerPhoneNumber", "SalesPerson",
]

# Pants header has PackagingMethod at the order level; Shirt does not
HEADER_COLS_PANTS = HEADER_COLS_COMMON + ["PackagingMethod"]
HEADER_COLS_SHIRT = HEADER_COLS_COMMON
HEADER_COLS       = HEADER_COLS_PANTS  # kept for backwards compat

HEADER_DETAIL_COLS = [
    "Courier", "CourierService", "ShipFromName",
    "ShipFromAdd1", "ShipFromAdd2", "ShipFromAdd3",
    "DCName", "DCNumber", "StoreNumber", "StoreName",
]

LINE_COLS_PANTS  = ["Product", "LineRefNo", "UnitRetailPrice", "Tier", "Quantity"]
LINE_COLS_SHIRT  = ["Product", "Part", "LineRefNo", "UnitRetailPrice", "Tier", "Quantity"]
LINE_COLS        = LINE_COLS_PANTS  # default; overridden per-row in excel_to_json

# Canonical order for Pants OrderLineDetail fields.
# LInseam short / RInseam short occupy the same position slot as LInseam / RInseam.
LINE_DETAIL_ORDER_PANTS = [
    "Style", "MainFabric", "StanttFabric", "FabricName", "PocketingFabric",
    "FrontStyle", "FrontCrease", "Hem", "FrenchFly", "WaistbandExtension",
    "WaistbandGripper", "Button", "Washing", "Dipping", "Monogram",
    "MonogramInitial", "MonogramLocation", "MonogramFont", "MonogramColor",
    "Size", "Fit", "Hip", "Waist",
    "LInseam", "RInseam",
    "LInseam short", "RInseam short",
    "IncrRise", "DecrRise", "Thigh", "Knee", "Ankle",
    "FrontPocket", "FrontPocketReinforcement", "BackPocket", "BackPocketLabel",
    "Label", "EDI_PO", "HangTag1", "HangTag2", "Rush",
    "StandardLabel1", "StandardLabel2", "StandardLabel3",
    # JokerTag injected here when StandardLabel3 has value
    "StandardLabel4", "PackingSlipSize", "Zipper", "Rivet",
    # UPC/ColorCode/ColorName/StyleName/StoreStyle injected here when StandardLabel3 has value
    "NewAlteration",
]

# Canonical order for Shirt OrderLineDetail fields.
LINE_DETAIL_ORDER_SHIRT = [
    "RushOrder", "Style", "PackagingMethod", "MainFabric", "StanttFabric", "FabricName",
    "Collar", "Cuff", "Placket", "CollarBandButton", "CollarPiping", "Hem",
    "Pocket", "PocketType", "ButtonColor", "CenterBackCollarButton",
    "ContrastButtonHoleStitch", "ContrastButtonHoleStitchLocation",
    "ButtonThreadColor", "TopButtonThreadColor", "Washing", "ArmHoleTaping",
    "Contrast", "InsideContrastID",
    "Monogram", "MonogramInitial", "MonogramLocation", "MonogramFont", "MonogramColor",
    "EmbroideryImage", "EmbroideryLocation", "EmbroideryStyle",
    "Grosgrain", "Yoke", "Pleats",
    "Size", "Tail", "Waist", "Neck", "Bicep", "LeftSleeve", "RightSleeve",
    "Hip", "AcrossShoulder", "Armhole",
    "LeftCuffHeight", "RightCuffHeight", "CollarRoll", "Chest",
    "LeftCuffCirc", "RightCuffCirc", "CollarStay",
    "MainLabel", "SizeLabel", "StandardLabel1", "StandardLabel2", "CareLabelText",
    "HangTag1", "HangTag2",
    "UPC", "UPC2", "UPC3", "QRURL", "EDI_PO", "SPI",
    "PackingSlipSize", "PackingSlipDesc",
    "HangerLoop", "HangerLoopContrast", "HangerLoopPlacement",
    "ColorCode", "ColorName", "StyleName", "StoreStyle",
    "NewAlteration",
]


def _pivot_detail(detail_list):
    """Convert [{ref, val}] list to a flat dict."""
    if not detail_list:
        return {}
    return {item.get("ref", ""): item.get("val", "") for item in detail_list}


def _is_shirt(line: dict) -> bool:
    return str(line.get("Product", "")).strip().lower() == "shirt"


def json_to_excel(json_bytes: bytes) -> bytes:
    data = json.loads(json_bytes)
    orders = data.get("Order", [])

    rows = []
    for order in orders:
        order_header = order.get("OrderHeader", order)
        # Detect product type from first line to pick correct header cols
        first_line = order.get("OrderLine", [{}])[0]
        hcols = HEADER_COLS_SHIRT if _is_shirt(first_line) else HEADER_COLS_PANTS
        header = {col: order_header.get(col, "") for col in hcols}
        header_detail = _pivot_detail(order_header.get("OrderHeaderDetail", []))

        for line in order.get("OrderLine", []):
            # Pick the right line-level columns based on product type
            lcols = LINE_COLS_SHIRT if _is_shirt(line) else LINE_COLS_PANTS
            line_fields = {col: line.get(col, "") for col in lcols}
            line_detail = _pivot_detail(line.get("OrderLineDetail", []))
            row = {**header, **header_detail, **line_fields, **line_detail}
            rows.append(row)

    if not rows:
        rows = [{}]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Orders"

    if rows:
        # Collect ALL columns across all rows to avoid missing fields
        # that only appear in some orders (e.g. JokerTag, LInseam short, etc.)
        seen = {}
        for row in rows:
            for k in row.keys():
                if k not in seen:
                    seen[k] = True
        columns = list(seen.keys())
        # Write header row
        header_fill = PatternFill("solid", fgColor="4472C4")
        header_font = Font(bold=True, color="FFFFFF")
        for col_idx, col_name in enumerate(columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        # Write data rows
        for row_idx, row in enumerate(rows, start=2):
            for col_idx, col_name in enumerate(columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

        # Auto-fit columns (approximate)
        for col_idx, col_name in enumerate(columns, start=1):
            max_len = max(len(str(col_name)), 10)
            for row_idx in range(2, min(len(rows) + 2, 102)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, min(len(str(val)), 40))
            ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_to_json(excel_bytes: bytes) -> bytes:
    df = pd.read_excel(io.BytesIO(excel_bytes), dtype=str, keep_default_na=False, na_values=[]).fillna("")

    all_cols = list(df.columns)

    # All non-header, non-line-level cols — used as fallback for extra fields
    all_line_detail_cols = [
        c for c in all_cols
        if c not in HEADER_COLS_PANTS and c not in HEADER_COLS_SHIRT
        and c not in HEADER_DETAIL_COLS
        and c not in LINE_COLS_PANTS and c not in LINE_COLS_SHIRT
    ]

    orders = []
    for order_number, group in df.groupby("OrderNumber", sort=False):
        first = group.iloc[0]

        # Detect product type from first row of this order group
        product_type = str(first.get("Product", "")).strip().lower()
        is_shirt_order = product_type == "shirt"
        hcols = HEADER_COLS_SHIRT if is_shirt_order else HEADER_COLS_PANTS

        order_header = {col: first[col] for col in hcols if col in df.columns}

        header_detail = [
            {"ref": col, "val": first[col]}
            for col in HEADER_DETAIL_COLS
            if col in df.columns
        ]

        order_lines = []
        for _, row in group.iterrows():
            product = str(row.get("Product", "")).strip().lower()
            is_shirt = product == "shirt"

            lcols = LINE_COLS_SHIRT if is_shirt else LINE_COLS_PANTS
            line_fields = {col: row[col] for col in lcols if col in df.columns}

            row_vals = {col: row[col] for col in df.columns}

            if is_shirt:
                # Shirt: emit fields in canonical shirt order, no inseam/JokerTag logic
                line_detail = []
                known_shirt = set(LINE_DETAIL_ORDER_SHIRT)
                for col in LINE_DETAIL_ORDER_SHIRT:
                    if col not in df.columns:
                        continue
                    line_detail.append({"ref": col, "val": row_vals[col]})
                # Append any extra columns not in shirt schema (future-proof)
                for col in all_line_detail_cols:
                    if col not in known_shirt and col not in set(LINE_COLS_SHIRT):
                        line_detail.append({"ref": col, "val": row_vals[col]})
            else:
                # Pants: inseam mutual exclusion + conditional JokerTag/retail fields
                has_regular_inseam = bool(str(row_vals.get("LInseam", "")).strip()) if "LInseam" in df.columns else False
                has_short_inseam   = bool(str(row_vals.get("LInseam short", "")).strip()) if "LInseam short" in df.columns else False
                std_label3_val     = str(row_vals.get("StandardLabel3", "")).strip() if "StandardLabel3" in df.columns else ""

                line_detail = []
                for col in LINE_DETAIL_ORDER_PANTS:
                    if col in ("LInseam", "RInseam") and not has_regular_inseam:
                        continue
                    if col in ("LInseam short", "RInseam short") and not has_short_inseam:
                        continue
                    if col not in df.columns:
                        continue
                    line_detail.append({"ref": col, "val": row_vals[col]})
                    if col == "StandardLabel3" and std_label3_val:
                        line_detail.append({"ref": "JokerTag", "val": std_label3_val})
                    if col == "Rivet" and std_label3_val:
                        for rf in ["UPC", "ColorCode", "ColorName", "StyleName", "StoreStyle"]:
                            line_detail.append({"ref": rf, "val": row_vals.get(rf, "")})

                # Append any extra columns not in pants schema (future-proof)
                known_pants = set(LINE_DETAIL_ORDER_PANTS) | {"JokerTag", "UPC", "ColorCode", "ColorName", "StyleName", "StoreStyle"}
                for col in all_line_detail_cols:
                    if col not in known_pants and col not in set(LINE_COLS_PANTS):
                        line_detail.append({"ref": col, "val": row_vals[col]})

            order_line = {**line_fields, "OrderLineDetail": line_detail}
            order_lines.append(order_line)

        order = {
            "OrderHeader": {
                **order_header,
                "OrderHeaderDetail": header_detail,
            },
            "OrderLine": order_lines,
        }
        orders.append(order)

    result = {"Order": orders}
    return json.dumps(result, indent=2, ensure_ascii=False).encode("utf-8")
