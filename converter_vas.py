import io
from xml.etree import ElementTree as ET
from xml.dom import minidom

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

# ── VAS (Vastrm) order XML <-> Excel ───────────────────────────────────────────
# XML shape:
#   <Orders>
#     <vastrm_grab_date>...</vastrm_grab_date>
#     <vastrm_run_dttm_utc>...</vastrm_run_dttm_utc>
#     <OrderHeader><Client>...</Client><OrderCount>...</OrderCount></OrderHeader>
#     <OrderDetail>
#       <vastrm_user_id>...</vastrm_user_id>
#       <vastrm_order_id>...</vastrm_order_id>
#       <vastrm_orderitem_id>...</vastrm_orderitem_id>
#       <vastrm_seq_num>...</vastrm_seq_num>
#       <vastrm_qty>...</vastrm_qty>
#       <Order><OrderId>...</OrderId><ReceiveDate>...</ReceiveDate><Qty>...</Qty></Order>
#       <Spec><ref>...</ref><val>...</val></Spec> * N
#     </OrderDetail> * N
#   </Orders>

DETAIL_ID_COLS = ["vastrm_user_id", "vastrm_order_id", "vastrm_orderitem_id", "vastrm_seq_num", "vastrm_qty"]
ORDER_COLS = ["OrderId", "ReceiveDate", "Qty"]

# File-level values, carried on every row as hidden columns so Excel -> XML
# round-trips without re-asking the user for them.
HIDDEN_HEADER_COLS = ["vastrm_grab_date", "vastrm_run_dttm_utc", "Client", "OrderCount"]

FIXED_COLS = HIDDEN_HEADER_COLS + DETAIL_ID_COLS + ORDER_COLS


def _text(el, tag, default=""):
    child = el.find(tag)
    if child is None or child.text is None:
        return default
    return child.text.strip()


def xml_to_excel_vas(xml_bytes: bytes) -> bytes:
    root = ET.fromstring(xml_bytes)

    grab_date = _text(root, "vastrm_grab_date")
    run_dttm = _text(root, "vastrm_run_dttm_utc")
    order_header = root.find("OrderHeader")
    client = _text(order_header, "Client") if order_header is not None else ""
    order_count = _text(order_header, "OrderCount") if order_header is not None else ""

    # Collect Spec ref columns in first-seen order across all OrderDetail blocks.
    spec_cols = []
    seen = set()
    details = root.findall("OrderDetail")
    for detail in details:
        for spec in detail.findall("Spec"):
            ref = _text(spec, "ref")
            if ref and ref not in seen:
                seen.add(ref)
                spec_cols.append(ref)

    columns = FIXED_COLS + spec_cols

    rows = []
    for detail in details:
        row = {
            "vastrm_grab_date": grab_date,
            "vastrm_run_dttm_utc": run_dttm,
            "Client": client,
            "OrderCount": order_count,
        }
        for col in DETAIL_ID_COLS:
            row[col] = _text(detail, col)

        order_el = detail.find("Order")
        if order_el is not None:
            for col in ORDER_COLS:
                row[col] = _text(order_el, col)
        else:
            for col in ORDER_COLS:
                row[col] = ""

        row_specs = {}
        for spec in detail.findall("Spec"):
            ref = _text(spec, "ref")
            val = _text(spec, "val")
            if ref:
                row_specs[ref] = val
        for col in spec_cols:
            row[col] = row_specs.get(col, "")

        rows.append(row)

    if not rows:
        rows = [{col: "" for col in columns}]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "VAS Orders"

    header_fill = PatternFill("solid", fgColor="4472C4")
    header_font = Font(bold=True, color="FFFFFF")
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row.get(col_name, ""))

    for col_idx, col_name in enumerate(columns, start=1):
        max_len = max(len(str(col_name)), 10)
        for row_idx in range(2, min(len(rows) + 2, 102)):
            val = ws.cell(row=row_idx, column=col_idx).value
            if val:
                max_len = max(max_len, min(len(str(val)), 40))
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = max_len + 2

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Hide the file-level columns that only exist to round-trip Excel -> XML.
    for col_name in HIDDEN_HEADER_COLS:
        col_idx = columns.index(col_name) + 1
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].hidden = True

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def read_excel_bytes_vas(excel_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(excel_bytes), data_only=True)
    ws = wb.active

    headers = [
        str(cell.value).strip() if cell.value is not None else ""
        for cell in next(ws.iter_rows(min_row=1, max_row=1))
    ]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(v is None for v in row):
            continue
        record = {headers[i]: v for i, v in enumerate(row) if i < len(headers) and headers[i]}
        rows.append(record)

    return rows


def _clean(val):
    if val is None:
        return ""
    return str(val).strip()


def excel_to_xml_vas(excel_bytes: bytes) -> str:
    records = read_excel_bytes_vas(excel_bytes)

    root = ET.Element("Orders")

    first = records[0] if records else {}
    ET.SubElement(root, "vastrm_grab_date").text = _clean(first.get("vastrm_grab_date"))
    ET.SubElement(root, "vastrm_run_dttm_utc").text = _clean(first.get("vastrm_run_dttm_utc"))

    order_header = ET.SubElement(root, "OrderHeader")
    ET.SubElement(order_header, "Client").text = _clean(first.get("Client"))
    ET.SubElement(order_header, "OrderCount").text = _clean(first.get("OrderCount")) or str(len(records))

    all_cols = set()
    for r in records:
        all_cols.update(r.keys())
    spec_cols = [c for c in all_cols if c not in FIXED_COLS]
    # Keep spec columns in a stable order: first-seen across records.
    ordered_spec_cols = []
    seen = set()
    for r in records:
        for c in r.keys():
            if c not in FIXED_COLS and c not in seen:
                seen.add(c)
                ordered_spec_cols.append(c)

    for record in records:
        detail = ET.SubElement(root, "OrderDetail")
        for col in DETAIL_ID_COLS:
            ET.SubElement(detail, col).text = _clean(record.get(col))

        order_el = ET.SubElement(detail, "Order")
        for col in ORDER_COLS:
            ET.SubElement(order_el, col).text = _clean(record.get(col))

        for col in ordered_spec_cols:
            if col not in record:
                continue
            spec = ET.SubElement(detail, "Spec")
            ET.SubElement(spec, "ref").text = col
            ET.SubElement(spec, "val").text = _clean(record.get(col))

    raw = ET.tostring(root, encoding="unicode")
    dom = minidom.parseString(raw)
    lines = dom.toprettyxml(indent="  ").splitlines()
    if lines[0].startswith("<?xml"):
        lines = lines[1:]
    xml_str = '<?xml version="1.0" encoding="UTF-8"?>\n' + "\n".join(line for line in lines if line.strip())

    return xml_str, len(records)
